"""Import investors from an Excel workbook (.xlsx).

Flow:
  1. Pick a file (or have one pre-filled via drag-drop).
  2. Pick the sheet.
  3. Map each app field to an Excel column using auto-match + override dropdowns.
  4. Preview the first handful of rows as they will be imported.
  5. WI% sum validation — warn if not 100% but allow proceeding.
  6. Click Import → rows created, per-row errors reported.

Uses openpyxl (already a runtime dep).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from PySide6.QtCore import Qt
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QPushButton,
    QRadioButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from wellsign.db.investors import insert_investor
from wellsign.db.projects import ProjectRow


# ---------------------------------------------------------------------------
# Field definitions
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class FieldDef:
    key: str
    label: str
    aliases: tuple[str, ...]
    required: bool = False


_FIELDS: list[FieldDef] = [
    FieldDef("first_name",   "First name",   ("first", "first name", "firstname", "fname", "given name")),
    FieldDef("last_name",    "Last name",    ("last", "last name", "lastname", "lname", "surname", "family name")),
    FieldDef("entity_name",  "Entity",       ("entity", "entity name", "company", "company name", "llc", "trust", "ira")),
    FieldDef("title",        "Title",        ("title", "position", "role")),
    FieldDef("email",        "Email",        ("email", "e-mail", "email address"), required=True),
    FieldDef("phone",        "Phone",        ("phone", "telephone", "cell", "mobile", "phone number")),
    FieldDef("address_line1","Address",      ("address", "address 1", "address1", "street", "address line 1", "addr")),
    FieldDef("city",         "City",         ("city", "town", "municipality")),
    FieldDef("state",        "State",        ("state", "province", "region")),
    FieldDef("zip_code",     "Zip",          ("zip", "zipcode", "zip code", "postal", "postal code", "postcode")),
    FieldDef("wi_percent",   "WI %",         ("wi", "wi%", "wi percent", "working interest", "interest", "percent", "%", "wi pct"), required=True),
    FieldDef("payment_preference", "Payment pref", ("payment", "payment method", "payment preference", "wire/check", "method")),
    FieldDef("notes",        "Notes",        ("notes", "comments", "remarks", "memo")),
]

_NONE_LABEL = "— none —"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _normalize(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace("_", " ").replace("-", " ").replace(".", "")


def _auto_match(headers: list[str], field: FieldDef) -> int:
    norm_headers = [_normalize(h) for h in headers]
    for alias in field.aliases:
        a = _normalize(alias)
        if a in norm_headers:
            return norm_headers.index(a)
    # Partial match fallback: alias is substring of header OR vice versa
    for alias in field.aliases:
        a = _normalize(alias)
        for i, h in enumerate(norm_headers):
            if a and h and (a in h or h in a):
                return i
    return -1


def _parse_wi(value: Any, percent_mode: bool) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        s = value.strip().rstrip("%").strip()
        try:
            v = float(s)
        except ValueError:
            return None
        if "%" in value:
            return v / 100.0
    else:
        try:
            v = float(value)
        except (TypeError, ValueError):
            return None
    return v / 100.0 if percent_mode else v


def _detect_percent_mode(column_values: list[Any]) -> bool:
    """Return True if numbers look like percentages (5.0 = 5%), False if decimals (0.05 = 5%)."""
    max_val = 0.0
    any_str_pct = False
    for v in column_values:
        if isinstance(v, str) and "%" in v:
            any_str_pct = True
            continue
        try:
            f = float(v)
        except (TypeError, ValueError):
            continue
        if f > max_val:
            max_val = f
    if any_str_pct:
        return True  # strings with % signs — input is percentage-style
    return max_val > 1.0


def _cell_value(c: Cell | Any) -> Any:
    if hasattr(c, "value"):
        return c.value
    return c


# ---------------------------------------------------------------------------
# Dialog
# ---------------------------------------------------------------------------
class ImportInvestorsDialog(QDialog):
    def __init__(
        self,
        project: ProjectRow,
        parent: QWidget | None = None,
        initial_file: Path | None = None,
    ) -> None:
        super().__init__(parent)
        self._project = project
        self._workbook = None
        self._headers: list[str] = []
        self._data_rows: list[list[Any]] = []
        self._mapping_combos: dict[str, QComboBox] = {}

        self.setWindowTitle("Import Investors from Excel")
        self.setModal(True)
        self.setMinimumWidth(900)
        self.setMinimumHeight(640)

        self._build()
        self._wire()

        if initial_file is not None:
            self._load_file(initial_file)
        self._update_import_enabled()

    # ---- build ----------------------------------------------------------
    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(22, 18, 22, 18)
        outer.setSpacing(12)

        title = QLabel("Import investors from Excel")
        f = title.font()
        f.setPointSize(14)
        f.setBold(True)
        title.setFont(f)
        outer.addWidget(title)

        subtitle = QLabel(
            "Pick an .xlsx file, confirm the column mapping, then preview the rows before "
            "committing. WI% must sum to 100%. Matching columns get auto-picked by name."
        )
        subtitle.setStyleSheet("color: #5b6473;")
        subtitle.setWordWrap(True)
        outer.addWidget(subtitle)

        # File + sheet row
        top_row = QHBoxLayout()
        top_row.setSpacing(8)
        self.file_edit = QLineEdit()
        self.file_edit.setReadOnly(True)
        self.file_edit.setPlaceholderText("No file selected")
        self.browse_btn = QPushButton("Browse…")
        self.browse_btn.setProperty("secondary", True)
        top_row.addWidget(QLabel("File:"))
        top_row.addWidget(self.file_edit, 1)
        top_row.addWidget(self.browse_btn)
        outer.addLayout(top_row)

        sheet_row = QHBoxLayout()
        sheet_row.setSpacing(8)
        sheet_row.addWidget(QLabel("Sheet:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(240)
        sheet_row.addWidget(self.sheet_combo)
        sheet_row.addStretch(1)
        outer.addLayout(sheet_row)

        # Mapping + WI format
        mapping_frame = QFrame()
        mapping_frame.setStyleSheet(
            "QFrame { background: #ffffff; border: 1px solid #d8dce3; border-radius: 6px; }"
        )
        mapping_layout = QVBoxLayout(mapping_frame)
        mapping_layout.setContentsMargins(16, 14, 16, 14)
        mapping_layout.setSpacing(8)

        mapping_title = QLabel("Column mapping")
        mf = mapping_title.font()
        mf.setBold(True)
        mapping_title.setFont(mf)
        mapping_layout.addWidget(mapping_title)

        form = QFormLayout()
        form.setHorizontalSpacing(14)
        form.setVerticalSpacing(6)
        for fd in _FIELDS:
            combo = QComboBox()
            combo.setMinimumWidth(260)
            combo.addItem(_NONE_LABEL, userData=-1)
            self._mapping_combos[fd.key] = combo
            label = fd.label + (" *" if fd.required else "")
            form.addRow(label + ":", combo)
        mapping_layout.addLayout(form)

        # WI format toggle
        wi_row = QHBoxLayout()
        wi_row.setSpacing(10)
        wi_row.addWidget(QLabel("WI % format:"))
        self.wi_decimal_radio = QRadioButton("Decimal (0.05 = 5%)")
        self.wi_percent_radio = QRadioButton("Percent number (5.0 = 5%)")
        self.wi_percent_radio.setChecked(True)
        self._wi_group = QButtonGroup(self)
        self._wi_group.addButton(self.wi_decimal_radio)
        self._wi_group.addButton(self.wi_percent_radio)
        wi_row.addWidget(self.wi_percent_radio)
        wi_row.addWidget(self.wi_decimal_radio)
        wi_row.addStretch(1)
        mapping_layout.addLayout(wi_row)

        outer.addWidget(mapping_frame)

        # Preview
        preview_title = QLabel("Preview")
        pf = preview_title.font()
        pf.setBold(True)
        preview_title.setFont(pf)
        outer.addWidget(preview_title)

        self.preview_table = QTableWidget(0, 0)
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.setAlternatingRowColors(False)
        self.preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.preview_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.preview_table.setMinimumHeight(180)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        outer.addWidget(self.preview_table, 1)

        # Status line
        self.status_label = QLabel("Pick a file to get started.")
        self.status_label.setStyleSheet("color: #5b6473;")
        self.status_label.setWordWrap(True)
        outer.addWidget(self.status_label)

        # Buttons
        self.buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Ok
        )
        self.import_btn = self.buttons.button(QDialogButtonBox.StandardButton.Ok)
        self.import_btn.setText("Import")
        self.cancel_btn = self.buttons.button(QDialogButtonBox.StandardButton.Cancel)
        self.cancel_btn.setProperty("secondary", True)
        outer.addWidget(self.buttons)

    def _wire(self) -> None:
        self.browse_btn.clicked.connect(self._on_browse)
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        for combo in self._mapping_combos.values():
            combo.currentIndexChanged.connect(self._on_mapping_changed)
        self.wi_decimal_radio.toggled.connect(self._on_mapping_changed)
        self.wi_percent_radio.toggled.connect(self._on_mapping_changed)
        self.buttons.accepted.connect(self._on_import)
        self.buttons.rejected.connect(self.reject)

    # ---- file loading ---------------------------------------------------
    def _on_browse(self) -> None:
        path_str, _ = QFileDialog.getOpenFileName(
            self, "Select Investor Spreadsheet", str(Path.cwd()), "Excel (*.xlsx *.xlsm)"
        )
        if not path_str:
            return
        self._load_file(Path(path_str))

    def _load_file(self, path: Path) -> None:
        try:
            wb = load_workbook(str(path), data_only=True)
        except Exception as e:
            self.status_label.setText(
                f"<span style='color:#d1242f;'>Could not open file: {e}</span>"
            )
            return
        self._workbook = wb
        self.file_edit.setText(str(path))
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for name in wb.sheetnames:
            self.sheet_combo.addItem(name)
        self.sheet_combo.blockSignals(False)
        self._on_sheet_changed(0)

    def _on_sheet_changed(self, _idx: int) -> None:
        if self._workbook is None or self.sheet_combo.currentIndex() < 0:
            return
        ws = self._workbook[self.sheet_combo.currentText()]
        raw_rows = list(ws.iter_rows(values_only=False))
        if not raw_rows:
            self.status_label.setText(
                "<span style='color:#d1242f;'>Sheet is empty.</span>"
            )
            return
        self._headers = [str(_cell_value(c) or "").strip() for c in raw_rows[0]]
        self._data_rows = [
            [_cell_value(c) for c in row] for row in raw_rows[1:]
        ]
        # Trim trailing blank rows
        while self._data_rows and all(
            v is None or str(v).strip() == "" for v in self._data_rows[-1]
        ):
            self._data_rows.pop()

        # Populate mapping combos
        for combo in self._mapping_combos.values():
            combo.blockSignals(True)
            combo.clear()
            combo.addItem(_NONE_LABEL, userData=-1)
            for i, h in enumerate(self._headers):
                label = h if h else f"(column {i + 1})"
                combo.addItem(label, userData=i)
            combo.blockSignals(False)

        # Auto-match
        for fd in _FIELDS:
            idx = _auto_match(self._headers, fd)
            combo_idx = idx + 1 if idx >= 0 else 0
            self._mapping_combos[fd.key].setCurrentIndex(combo_idx)

        # Auto-detect WI format from the mapped WI column
        wi_col = self._mapping_combos["wi_percent"].currentData()
        if isinstance(wi_col, int) and wi_col >= 0:
            col_values = [row[wi_col] for row in self._data_rows if wi_col < len(row)]
            if _detect_percent_mode(col_values):
                self.wi_percent_radio.setChecked(True)
            else:
                self.wi_decimal_radio.setChecked(True)

        self._refresh_preview()
        self._update_import_enabled()

    # ---- mapping + preview ----------------------------------------------
    def _on_mapping_changed(self) -> None:
        self._refresh_preview()
        self._update_import_enabled()

    def _mapped_col(self, key: str) -> int:
        combo = self._mapping_combos.get(key)
        if combo is None:
            return -1
        data = combo.currentData()
        return int(data) if data is not None else -1

    def _refresh_preview(self) -> None:
        if not self._headers or not self._data_rows:
            self.preview_table.clear()
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            return

        show_fields = [
            fd for fd in _FIELDS
            if self._mapped_col(fd.key) >= 0
        ]
        if not show_fields:
            show_fields = _FIELDS  # show them all anyway so the user sees structure

        self.preview_table.clear()
        self.preview_table.setColumnCount(len(show_fields))
        self.preview_table.setHorizontalHeaderLabels([fd.label for fd in show_fields])

        preview_count = min(len(self._data_rows), 8)
        self.preview_table.setRowCount(preview_count)

        percent_mode = self.wi_percent_radio.isChecked()
        for r in range(preview_count):
            row = self._data_rows[r]
            for c, fd in enumerate(show_fields):
                col_idx = self._mapped_col(fd.key)
                if col_idx < 0 or col_idx >= len(row):
                    text = ""
                else:
                    raw = row[col_idx]
                    if fd.key == "wi_percent":
                        parsed = _parse_wi(raw, percent_mode)
                        text = f"{parsed * 100:.6f}%" if parsed is not None else "—"
                    elif raw is None:
                        text = ""
                    else:
                        text = str(raw)
                item = QTableWidgetItem(text)
                self.preview_table.setItem(r, c, item)

        # WI sum validation
        wi_col = self._mapped_col("wi_percent")
        if wi_col >= 0:
            total = 0.0
            for row in self._data_rows:
                if wi_col < len(row):
                    parsed = _parse_wi(row[wi_col], percent_mode)
                    if parsed is not None:
                        total += parsed
            pct = total * 100
            if abs(pct - 100.0) < 0.001:
                self.status_label.setText(
                    f"<span style='color:#1a7f37;'>"
                    f"<b>{len(self._data_rows)}</b> rows  ·  "
                    f"WI% sum: <b>{pct:.6f}%</b> ✓</span>"
                )
            else:
                self.status_label.setText(
                    f"<span style='color:#d97706;'>"
                    f"<b>{len(self._data_rows)}</b> rows  ·  "
                    f"WI% sum: <b>{pct:.6f}%</b> (should be 100% — import anyway is allowed)</span>"
                )
        else:
            self.status_label.setText(
                f"<span style='color:#d1242f;'>"
                f"Map the WI% column to continue.</span>"
            )

    def _update_import_enabled(self) -> None:
        has_data = bool(self._data_rows)
        has_email = self._mapped_col("email") >= 0
        has_wi = self._mapped_col("wi_percent") >= 0
        self.import_btn.setEnabled(has_data and has_email and has_wi)

    # ---- import ---------------------------------------------------------
    def _on_import(self) -> None:
        if self._project is None or not self._data_rows:
            return

        percent_mode = self.wi_percent_radio.isChecked()
        cols = {fd.key: self._mapped_col(fd.key) for fd in _FIELDS}

        total_llg = 0.0
        total_dhc = 0.0
        # Pull project totals if available for per-investor $ math
        from wellsign.db.migrate import connect
        with connect() as conn:
            row = conn.execute(
                "SELECT total_llg_cost, total_dhc_cost FROM projects WHERE id = ?",
                (self._project.id,),
            ).fetchone()
            if row:
                total_llg = float(row["total_llg_cost"] or 0)
                total_dhc = float(row["total_dhc_cost"] or 0)

        imported = 0
        skipped = 0
        errors: list[str] = []

        for r, row in enumerate(self._data_rows):
            def _get(key: str) -> Any:
                idx = cols.get(key, -1)
                if idx < 0 or idx >= len(row):
                    return None
                val = row[idx]
                return val if val not in (None, "") else None

            email_val = _get("email")
            if not email_val:
                skipped += 1
                errors.append(f"Row {r + 2}: no email — skipped")
                continue

            wi_val = _parse_wi(_get("wi_percent"), percent_mode)
            if wi_val is None:
                skipped += 1
                errors.append(f"Row {r + 2}: could not parse WI% — skipped")
                continue

            llg_amt = round(wi_val * total_llg, 2) if total_llg else None
            dhc_amt = round(wi_val * total_dhc, 2) if total_dhc else None

            def _str(key: str) -> str | None:
                v = _get(key)
                return str(v).strip() if v is not None else None

            try:
                insert_investor(
                    project_id=self._project.id,
                    first_name=_str("first_name"),
                    last_name=_str("last_name"),
                    entity_name=_str("entity_name"),
                    title=_str("title"),
                    email=str(email_val).strip(),
                    phone=_str("phone"),
                    address_line1=_str("address_line1"),
                    city=_str("city"),
                    state=_str("state"),
                    zip_code=_str("zip_code"),
                    wi_percent=wi_val,
                    llg_amount=llg_amt,
                    dhc_amount=dhc_amt,
                    payment_preference=_str("payment_preference"),
                    notes=_str("notes"),
                )
                imported += 1
            except Exception as e:  # noqa: BLE001
                skipped += 1
                errors.append(f"Row {r + 2}: {e}")

        # Stash results for the caller
        self._imported_count = imported
        self._skipped_count = skipped
        self._errors = errors
        self.accept()

    @property
    def imported_count(self) -> int:
        return getattr(self, "_imported_count", 0)

    @property
    def skipped_count(self) -> int:
        return getattr(self, "_skipped_count", 0)

    @property
    def errors(self) -> list[str]:
        return getattr(self, "_errors", [])
